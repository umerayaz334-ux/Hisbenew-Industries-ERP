from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
import base64
import hashlib
import hmac
import os
from pathlib import Path
import re
import secrets
import time
from typing import Any
from urllib.parse import urlparse
from urllib.request import Request, urlopen
import uuid
import zipfile

from PIL import Image as PilImage, ImageDraw, ImageFont, ImageOps
from sqlalchemy.orm import Session

from .config import SECRET_KEY
from .models import Product, StockMovement


MAX_FAIRE_WORKBOOK_BYTES = 30 * 1024 * 1024
MAX_FAIRE_IMAGE_BYTES = 12 * 1024 * 1024
MAX_FAIRE_ROWS = 10_000
FAIRE_IMAGE_WORKERS = 6
CATALOG_DOWNLOAD_TOKEN_TTL_SECONDS = 5 * 60


class ProductCatalogError(ValueError):
    pass


def create_catalog_download_token(
    user_id: int,
    product_ids: list[int] | None = None,
    ttl_seconds: int = CATALOG_DOWNLOAD_TOKEN_TTL_SECONDS,
) -> str:
    expires_at = int(time.time()) + max(int(ttl_seconds), 1)
    selected = "" if product_ids is None else ",".join(str(int(value)) for value in sorted(set(product_ids)))
    payload = f"catalog:{int(user_id)}:{expires_at}:{secrets.token_urlsafe(12)}:{selected}"
    encoded_payload = base64.urlsafe_b64encode(payload.encode("utf-8")).decode("ascii").rstrip("=")
    signature = hmac.new(
        SECRET_KEY.encode("utf-8"),
        encoded_payload.encode("ascii"),
        hashlib.sha256,
    ).digest()
    encoded_signature = base64.urlsafe_b64encode(signature).decode("ascii").rstrip("=")
    return f"{encoded_payload}.{encoded_signature}"


def decode_catalog_download_token(token: str) -> tuple[int, list[int] | None]:
    try:
        encoded_payload, encoded_signature = str(token or "").split(".", 1)
        expected_signature = hmac.new(
            SECRET_KEY.encode("utf-8"),
            encoded_payload.encode("ascii"),
            hashlib.sha256,
        ).digest()
        supplied_signature = base64.urlsafe_b64decode(
            encoded_signature + "=" * (-len(encoded_signature) % 4)
        )
        if not hmac.compare_digest(expected_signature, supplied_signature):
            raise ProductCatalogError("The catalog download link is invalid.")
        payload = base64.urlsafe_b64decode(
            encoded_payload + "=" * (-len(encoded_payload) % 4)
        ).decode("utf-8")
        parts = payload.split(":", 4)
        if len(parts) == 4:
            scope, user_id, expires_at, _nonce = parts
            selected_ids = None
        else:
            scope, user_id, expires_at, _nonce, selected = parts
            selected_ids = [int(value) for value in selected.split(",") if value]
        if scope != "catalog" or int(expires_at) < int(time.time()):
            raise ProductCatalogError("The catalog download link has expired.")
        return int(user_id), selected_ids
    except ProductCatalogError:
        raise
    except Exception as exc:
        raise ProductCatalogError("The catalog download link is invalid.") from exc


def verify_catalog_download_token(token: str) -> int:
    return decode_catalog_download_token(token)[0]


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_decimal(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float, Decimal)):
        return max(float(value), 0.0)
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    if not cleaned or cleaned in {"-", ".", "-."}:
        return 0.0
    try:
        return max(float(Decimal(cleaned)), 0.0)
    except (InvalidOperation, ValueError):
        return 0.0


def _parse_inventory(value: Any) -> int:
    return max(int(_parse_decimal(value)), 0)


def _weight_in_kg(value: Any, unit: Any) -> float:
    weight = _parse_decimal(value)
    normalized_unit = _clean_text(unit).lower()
    multiplier = {
        "kg": 1.0,
        "g": 0.001,
        "lb": 0.45359237,
        "oz": 0.028349523125,
    }.get(normalized_unit, 0.0)
    return round(weight * multiplier, 6) if multiplier else 0.0


def _first_http_url(value: Any) -> str | None:
    match = re.search(r"https?://[^\s]+", _clean_text(value), flags=re.IGNORECASE)
    return match.group(0).rstrip(",;") if match else None


def _faire_image_url(value: Any) -> str | None:
    url = _first_http_url(value)
    if not url:
        return None
    parsed = urlparse(url)
    hostname = (parsed.hostname or "").lower()
    if parsed.scheme not in {"http", "https"}:
        return None
    if hostname != "faire.com" and not hostname.endswith(".faire.com"):
        return None
    return url


def _option_summary(row: dict[str, Any]) -> str | None:
    options: list[str] = []
    for index in range(1, 4):
        name = _clean_text(row.get(f"option_{index}_name"))
        value = _clean_text(row.get(f"option_{index}_value"))
        if name and value:
            options.append(f"{name}: {value}")
        elif value:
            options.append(value)
    return "; ".join(options) or None


def parse_faire_workbook(content: bytes) -> dict[str, Any]:
    if not content:
        raise ProductCatalogError("The Faire workbook is empty.")
    if len(content) > MAX_FAIRE_WORKBOOK_BYTES:
        raise ProductCatalogError("The Faire workbook is larger than 30 MB.")
    if not zipfile.is_zipfile(BytesIO(content)):
        raise ProductCatalogError("Upload the original Faire .xlsx workbook.")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise ProductCatalogError("Spreadsheet import support is not installed.") from exc

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ProductCatalogError("The Faire workbook could not be opened.") from exc

    try:
        sheet = next(
            (item for item in workbook.worksheets if item.title.strip().lower() == "products"),
            None,
        )
        if sheet is None:
            raise ProductCatalogError("The workbook does not contain a Products sheet.")
        if sheet.max_row > MAX_FAIRE_ROWS:
            raise ProductCatalogError("The Faire workbook contains too many product rows.")

        machine_header_row = None
        header_indexes: dict[str, int] = {}
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 12), values_only=True),
            start=1,
        ):
            candidate = {
                _clean_text(value).lower(): index
                for index, value in enumerate(values)
                if _clean_text(value)
            }
            if {"product_name_english", "sku", "price_wholesale", "price_retail"}.issubset(
                candidate
            ):
                machine_header_row = row_number
                header_indexes = candidate
                break

        if machine_header_row is None:
            raise ProductCatalogError(
                "The Products sheet is not a supported Faire bulk export."
            )

        products: list[dict[str, Any]] = []
        skipped_rows: list[dict[str, Any]] = []
        seen_skus: set[str] = set()
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=machine_header_row + 1, values_only=True),
            start=machine_header_row + 1,
        ):
            row = {
                header: values[index] if index < len(values) else None
                for header, index in header_indexes.items()
            }
            if not any(value not in (None, "") for value in row.values()):
                continue

            sku = _clean_text(row.get("sku"))
            name = _clean_text(row.get("product_name_english"))
            product_status = _clean_text(row.get("info_status_v2")).lower()
            option_status = _clean_text(row.get("option_status")).lower()
            if product_status == "deleted" or option_status == "deleted":
                skipped_rows.append({"row": row_number, "reason": "Deleted on Faire"})
                continue
            if not sku:
                skipped_rows.append({"row": row_number, "reason": "SKU is missing"})
                continue
            sku_key = sku.casefold()
            if sku_key in seen_skus:
                skipped_rows.append({"row": row_number, "reason": f"Duplicate SKU {sku}"})
                continue
            seen_skus.add(sku_key)

            image_source = _faire_image_url(row.get("option_image")) or _faire_image_url(
                row.get("product_images")
            )
            products.append(
                {
                    "row": row_number,
                    "sku": sku,
                    "name": name or sku,
                    "category": _clean_text(row.get("info_product_type")) or "Uncategorized",
                    "description": _clean_text(row.get("product_description_english")),
                    "options": _option_summary(row),
                    "wholesale_price": _parse_decimal(row.get("price_wholesale")),
                    "retail_price": _parse_decimal(row.get("price_retail")),
                    "unit_weight_kg": _weight_in_kg(
                        row.get("item_weight"), row.get("item_weight_unit")
                    ),
                    "inventory": _parse_inventory(row.get("on_hand_inventory")),
                    "image_source": image_source,
                }
            )

        if not products:
            raise ProductCatalogError("No importable products with SKUs were found.")
        return {
            "products": products,
            "source_rows": len(products) + len(skipped_rows),
            "skipped_rows": skipped_rows,
        }
    finally:
        workbook.close()


def _cached_faire_image(url: str, upload_dir: Path) -> str:
    safe_url = _faire_image_url(url)
    if not safe_url:
        raise ProductCatalogError("Only Faire-hosted product images can be imported.")

    upload_dir.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha256(safe_url.encode("utf-8")).hexdigest()[:32]
    existing = next(upload_dir.glob(f"faire-{digest}.*"), None)
    if existing and existing.is_file():
        return f"/static/uploads/{existing.name}"

    request = Request(
        safe_url,
        headers={
            "User-Agent": "Hisbenew-ERP/1.0 catalog-import",
            "Accept": "image/avif,image/webp,image/png,image/jpeg,*/*;q=0.5",
        },
    )
    with urlopen(request, timeout=15) as response:
        content_length = response.headers.get("Content-Length")
        if content_length and int(content_length) > MAX_FAIRE_IMAGE_BYTES:
            raise ProductCatalogError("A Faire image is larger than 12 MB.")
        content = response.read(MAX_FAIRE_IMAGE_BYTES + 1)
    if len(content) > MAX_FAIRE_IMAGE_BYTES:
        raise ProductCatalogError("A Faire image is larger than 12 MB.")

    try:
        with PilImage.open(BytesIO(content)) as opened:
            opened.load()
            image = ImageOps.exif_transpose(opened)
            image.thumbnail((2400, 2400), PilImage.Resampling.LANCZOS)
            has_alpha = image.mode in {"RGBA", "LA"} or (
                image.mode == "P" and "transparency" in image.info
            )
            if has_alpha:
                output_format = "PNG"
                extension = ".png"
                prepared = image.convert("RGBA")
            else:
                output_format = "JPEG"
                extension = ".jpg"
                prepared = image.convert("RGB")
            final_path = upload_dir / f"faire-{digest}{extension}"
            temporary_path = upload_dir / f"faire-{digest}-{uuid.uuid4().hex}.tmp"
            try:
                save_options = {"optimize": True}
                if output_format == "JPEG":
                    save_options["quality"] = 90
                prepared.save(temporary_path, format=output_format, **save_options)
                os.replace(temporary_path, final_path)
            finally:
                if temporary_path.exists():
                    temporary_path.unlink()
    except ProductCatalogError:
        raise
    except Exception as exc:
        raise ProductCatalogError("A Faire product image is not valid.") from exc

    return f"/static/uploads/{final_path.name}"


def cache_faire_product_images(
    rows: list[dict[str, Any]], upload_dir: Path
) -> tuple[dict[str, str], list[dict[str, str]]]:
    urls = sorted({row["image_source"] for row in rows if row.get("image_source")})
    if not urls:
        return {}, []

    cached: dict[str, str] = {}
    failures: list[dict[str, str]] = []
    with ThreadPoolExecutor(max_workers=FAIRE_IMAGE_WORKERS) as executor:
        future_urls = {
            executor.submit(_cached_faire_image, url, upload_dir): url for url in urls
        }
        for future in as_completed(future_urls):
            url = future_urls[future]
            try:
                cached[url] = future.result()
            except Exception as exc:
                failures.append({"url": url, "reason": str(exc)})
    return cached, failures


def import_faire_products(
    db: Session,
    parsed: dict[str, Any],
    upload_dir: Path,
) -> dict[str, Any]:
    rows = list(parsed["products"])
    cached_images, image_failures = cache_faire_product_images(rows, upload_dir)
    existing_products = {
        product.article_no.strip().casefold(): product
        for product in db.query(Product).all()
        if product.article_no and product.article_no.strip()
    }

    created = 0
    updated = 0
    try:
        for row in rows:
            sku_key = row["sku"].casefold()
            product = existing_products.get(sku_key)
            imported_image = cached_images.get(row.get("image_source"))
            effective_image = imported_image or row.get("image_source")
            if product is None:
                product = Product(
                    article_no=row["sku"],
                    name=row["name"],
                    category=row["category"],
                    image_url=effective_image,
                    options=row.get("options"),
                    notes=row.get("description") or None,
                    factory_stock=0,
                    usa_stock=row["inventory"],
                    front_room_stock=0,
                    reserved_stock=0,
                    cost_price=row["wholesale_price"],
                    selling_price=row["retail_price"],
                    unit_weight_kg=row["unit_weight_kg"],
                    low_stock_alert=10,
                    workflow_required=True,
                )
                db.add(product)
                db.flush()
                if row["inventory"] > 0:
                    db.add(
                        StockMovement(
                            product_id=product.id,
                            movement_type="Initial Stock",
                            quantity=row["inventory"],
                            stock_type="usa_stock",
                            source="Faire Catalog",
                            reference=product.article_no,
                            note="Opening USA stock imported from Faire on-hand inventory.",
                            created_at=datetime.utcnow(),
                        )
                    )
                existing_products[sku_key] = product
                created += 1
                continue

            product.name = row["name"]
            product.category = row["category"]
            product.options = row.get("options")
            product.cost_price = row["wholesale_price"]
            product.selling_price = row["retail_price"]
            product.unit_weight_kg = row["unit_weight_kg"]
            if effective_image:
                product.image_url = effective_image
            if not _clean_text(product.notes) and row.get("description"):
                product.notes = row["description"]
            db.add(product)
            updated += 1
        db.commit()
    except Exception:
        db.rollback()
        raise

    skipped_rows = list(parsed.get("skipped_rows") or [])
    return {
        "message": "Faire catalog imported successfully.",
        "source_rows": parsed.get("source_rows", len(rows)),
        "imported": len(rows),
        "created": created,
        "updated": updated,
        "skipped": len(skipped_rows),
        "skipped_rows": skipped_rows[:25],
        "images_cached": len(cached_images),
        "image_failures": len(image_failures),
        "image_failure_details": image_failures[:10],
        "existing_inventory_preserved": True,
    }


def import_faire_workbook(
    db: Session,
    content: bytes,
    upload_dir: Path,
) -> dict[str, Any]:
    return import_faire_products(db, parse_faire_workbook(content), upload_dir)


def _local_product_image_path(image_url: str | None, upload_dir: Path) -> Path | None:
    value = _clean_text(image_url)
    if not value or value.startswith("http://") or value.startswith("https://"):
        return None
    prefix = "/static/uploads/"
    if not value.startswith(prefix):
        return None
    filename = value[len(prefix) :]
    if not filename or Path(filename).name != filename:
        return None
    candidate = (upload_dir / filename).resolve()
    try:
        candidate.relative_to(upload_dir.resolve())
    except ValueError:
        return None
    return candidate if candidate.is_file() else None


def _catalog_thumbnail(image_url: str | None, upload_dir: Path) -> BytesIO:
    source_bytes: bytes | None = None
    local_path = _local_product_image_path(image_url, upload_dir)
    if local_path:
        try:
            source_bytes = local_path.read_bytes()
        except OSError:
            source_bytes = None
    elif _faire_image_url(image_url):
        try:
            cached_url = _cached_faire_image(str(image_url), upload_dir)
            cached_path = _local_product_image_path(cached_url, upload_dir)
            source_bytes = cached_path.read_bytes() if cached_path else None
        except Exception:
            source_bytes = None

    canvas_size = 720
    canvas = PilImage.new("RGB", (canvas_size, canvas_size), "white")
    if source_bytes:
        try:
            with PilImage.open(BytesIO(source_bytes)) as opened:
                opened.load()
                product_image = ImageOps.exif_transpose(opened).convert("RGBA")
                product_image.thumbnail((660, 660), PilImage.Resampling.LANCZOS)
                left = (canvas_size - product_image.width) // 2
                top = (canvas_size - product_image.height) // 2
                canvas.paste(product_image, (left, top), product_image)
        except Exception:
            source_bytes = None
    if not source_bytes:
        draw = ImageDraw.Draw(canvas)
        draw.rounded_rectangle((90, 90, 630, 630), radius=32, outline="#D7DCE3", width=8)
        font = ImageFont.load_default(size=32)
        label = "NO PRODUCT IMAGE"
        box = draw.textbbox((0, 0), label, font=font)
        draw.text(
            ((canvas_size - (box[2] - box[0])) / 2, (canvas_size - (box[3] - box[1])) / 2),
            label,
            fill="#7B8491",
            font=font,
        )

    output = BytesIO()
    canvas.save(output, format="JPEG", quality=88, optimize=True)
    output.seek(0)
    return output


def _shorten(value: Any, limit: int) -> str:
    text = re.sub(r"\s+", " ", _clean_text(value))
    return text if len(text) <= limit else f"{text[: max(0, limit - 3)].rstrip()}..."


def _usd(value: Any) -> str:
    return f"${float(value or 0):,.2f}"


def _build_product_catalog_pdf_flowable(products: list[Product], upload_dir: Path) -> bytes:
    if not products:
        raise ProductCatalogError("Add at least one product before downloading a catalog.")

    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        Image,
        KeepTogether,
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )
    from xml.sax.saxutils import escape

    obsidian = colors.HexColor("#0B0C0E")
    charcoal = colors.HexColor("#17191D")
    gold = colors.HexColor("#C79A43")
    bright_gold = colors.HexColor("#E1BC6A")
    champagne = colors.HexColor("#E8D7AF")
    ivory = colors.HexColor("#F6F2E9")
    paper = colors.HexColor("#FCFAF5")
    ink = colors.HexColor("#17191C")
    muted = colors.HexColor("#6D6A64")
    line = colors.HexColor("#DDD5C7")
    soft = colors.HexColor("#EEE8DC")
    brand_logo_path = (
        Path(__file__).resolve().parent / "assets" / "hisbenew-catalog-logo.png"
    )

    output = BytesIO()
    styles = getSampleStyleSheet()
    eyebrow_style = ParagraphStyle(
        "CatalogEyebrow",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.5,
        leading=11,
        textColor=bright_gold,
        alignment=TA_CENTER,
        spaceAfter=9,
    )
    cover_title_style = ParagraphStyle(
        "CatalogCoverTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=29,
        leading=31,
        textColor=colors.white,
        alignment=TA_CENTER,
        spaceAfter=8,
    )
    cover_subtitle_style = ParagraphStyle(
        "CatalogCoverSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10.5,
        leading=15,
        textColor=champagne,
        alignment=TA_CENTER,
        leftIndent=0.5 * inch,
        rightIndent=0.5 * inch,
        spaceAfter=17,
    )
    category_style = ParagraphStyle(
        "CatalogCategory",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12.5,
        leading=15,
        textColor=colors.white,
        spaceAfter=0,
    )
    category_number_style = ParagraphStyle(
        "CatalogCategoryNumber",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=15,
        leading=18,
        textColor=obsidian,
        alignment=TA_CENTER,
    )
    category_count_style = ParagraphStyle(
        "CatalogCategoryCount",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=7.2,
        leading=9,
        textColor=champagne,
        alignment=TA_RIGHT,
    )
    name_style = ParagraphStyle(
        "CatalogProductName",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9.1,
        leading=11,
        textColor=ink,
        alignment=TA_LEFT,
    )
    sku_style = ParagraphStyle(
        "CatalogSku",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=7.1,
        leading=8.5,
        textColor=gold,
    )
    price_label_style = ParagraphStyle(
        "CatalogPriceLabel",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=6.2,
        leading=7.5,
        textColor=champagne,
    )
    wholesale_style = ParagraphStyle(
        "CatalogWholesale",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=11.5,
        leading=13,
        textColor=bright_gold,
    )
    msrp_style = ParagraphStyle(
        "CatalogMsrp",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9.5,
        leading=13,
        textColor=colors.white,
        alignment=TA_RIGHT,
    )

    def page_decoration(canvas, doc):
        canvas.saveState()
        width, height = letter
        if doc.page == 1:
            canvas.setFillColor(obsidian)
            canvas.rect(0, 0, width, height, stroke=0, fill=1)

            canvas.setFillColor(charcoal)
            canvas.circle(width + 0.2 * inch, height - 1.15 * inch, 1.8 * inch, stroke=0, fill=1)
            canvas.circle(-0.25 * inch, 1.0 * inch, 1.35 * inch, stroke=0, fill=1)

            canvas.setStrokeColor(gold)
            canvas.setLineWidth(0.8)
            canvas.rect(0.24 * inch, 0.24 * inch, width - 0.48 * inch, height - 0.48 * inch, stroke=1, fill=0)
            canvas.setStrokeColor(colors.HexColor("#5B4825"))
            canvas.setLineWidth(0.35)
            canvas.rect(0.31 * inch, 0.31 * inch, width - 0.62 * inch, height - 0.62 * inch, stroke=1, fill=0)

            canvas.setFillColor(gold)
            canvas.rect(0.24 * inch, 0.24 * inch, 0.08 * inch, height - 0.48 * inch, stroke=0, fill=1)
            canvas.setFont("Helvetica-Bold", 6.8)
            canvas.setFillColor(champagne)
            canvas.drawString(0.5 * inch, 0.43 * inch, "HISBENEW INDUSTRIES")
            canvas.setFillColor(bright_gold)
            canvas.drawRightString(width - 0.5 * inch, 0.43 * inch, "TRADE PARTNER EDITION")
        else:
            canvas.setFillColor(ivory)
            canvas.rect(0, 0, width, height, stroke=0, fill=1)
            canvas.setFillColor(obsidian)
            canvas.rect(0, height - 0.54 * inch, width, 0.54 * inch, stroke=0, fill=1)
            canvas.setFillColor(gold)
            canvas.rect(0, height - 0.575 * inch, width, 0.035 * inch, stroke=0, fill=1)

            canvas.setFont("Helvetica-Bold", 7.4)
            canvas.setFillColor(bright_gold)
            canvas.drawString(doc.leftMargin, height - 0.34 * inch, "HISBENEW")
            canvas.setFont("Helvetica", 7.1)
            canvas.setFillColor(champagne)
            canvas.drawRightString(
                width - doc.rightMargin,
                height - 0.34 * inch,
                "WHOLESALE COLLECTION  |  USD",
            )

            canvas.setStrokeColor(line)
            canvas.setLineWidth(0.45)
            canvas.line(doc.leftMargin, 0.43 * inch, width - doc.rightMargin, 0.43 * inch)
            canvas.setFont("Helvetica", 6.9)
            canvas.setFillColor(muted)
            canvas.drawString(doc.leftMargin, 0.25 * inch, "WHOLESALE AND MSRP SHOWN IN US DOLLARS")
            canvas.setFont("Helvetica-Bold", 7)
            canvas.setFillColor(ink)
            canvas.drawRightString(width - doc.rightMargin, 0.25 * inch, f"{doc.page:02d}")
        canvas.restoreState()

    def page_chrome(canvas, doc):
        """Redraw page furniture after flowables so long tables cannot cover it."""
        canvas.saveState()
        width, height = letter
        if doc.page == 1:
            canvas.setStrokeColor(gold)
            canvas.setLineWidth(0.8)
            canvas.rect(0.24 * inch, 0.24 * inch, width - 0.48 * inch, height - 0.48 * inch, stroke=1, fill=0)
            canvas.setStrokeColor(colors.HexColor("#5B4825"))
            canvas.setLineWidth(0.35)
            canvas.rect(0.31 * inch, 0.31 * inch, width - 0.62 * inch, height - 0.62 * inch, stroke=1, fill=0)
            canvas.setFillColor(gold)
            canvas.rect(0.24 * inch, 0.24 * inch, 0.08 * inch, height - 0.48 * inch, stroke=0, fill=1)
            canvas.setFont("Helvetica-Bold", 6.8)
            canvas.setFillColor(champagne)
            canvas.drawString(0.5 * inch, 0.43 * inch, "HISBENEW INDUSTRIES")
            canvas.setFillColor(bright_gold)
            canvas.drawRightString(width - 0.5 * inch, 0.43 * inch, "TRADE PARTNER EDITION")
        else:
            canvas.setFillColor(obsidian)
            canvas.rect(0, height - 0.54 * inch, width, 0.54 * inch, stroke=0, fill=1)
            canvas.setFillColor(gold)
            canvas.rect(0, height - 0.575 * inch, width, 0.035 * inch, stroke=0, fill=1)
            canvas.setFont("Helvetica-Bold", 7.4)
            canvas.setFillColor(bright_gold)
            canvas.drawString(doc.leftMargin, height - 0.34 * inch, "HISBENEW")
            canvas.setFont("Helvetica", 7.1)
            canvas.setFillColor(champagne)
            canvas.drawRightString(width - doc.rightMargin, height - 0.34 * inch, "WHOLESALE COLLECTION  |  USD")

            canvas.setStrokeColor(line)
            canvas.setLineWidth(0.45)
            canvas.line(doc.leftMargin, 0.43 * inch, width - doc.rightMargin, 0.43 * inch)
            canvas.setFont("Helvetica", 6.9)
            canvas.setFillColor(muted)
            canvas.drawString(doc.leftMargin, 0.25 * inch, "WHOLESALE AND MSRP SHOWN IN US DOLLARS")
            canvas.setFont("Helvetica-Bold", 7)
            canvas.setFillColor(ink)
            canvas.drawRightString(width - doc.rightMargin, 0.25 * inch, f"{doc.page:02d}")
        canvas.restoreState()

    class CatalogDocTemplate(SimpleDocTemplate):
        def afterPage(self):
            page_chrome(self.canv, self)

    document = CatalogDocTemplate(
        output,
        pagesize=letter,
        leftMargin=0.44 * inch,
        rightMargin=0.44 * inch,
        topMargin=0.76 * inch,
        bottomMargin=0.62 * inch,
        title="Hisbenew Industries Wholesale Product Catalog",
        author="Hisbenew Industries",
        subject="USD wholesale and MSRP product catalog",
    )

    grouped: dict[str, list[Product]] = {}
    for product in sorted(
        products,
        key=lambda item: (
            (_clean_text(item.category) or "Uncategorized").casefold(),
            _clean_text(item.name).casefold(),
            _clean_text(item.article_no).casefold(),
        ),
    ):
        grouped.setdefault(_clean_text(product.category) or "Uncategorized", []).append(product)

    story: list[Any] = [Spacer(1, 0.12 * inch)]
    if brand_logo_path.is_file():
        logo = Image(str(brand_logo_path), width=6.35 * inch, height=(6.35 / 1.5) * inch)
        logo.hAlign = "CENTER"
        story.extend([logo, Spacer(1, 0.05 * inch)])
    else:
        story.extend(
            [
                Spacer(1, 1.8 * inch),
                Paragraph("HISBENEW", cover_title_style),
                Spacer(1, 1.5 * inch),
            ]
        )
    story.extend(
        [
            Paragraph("2026 TRADE CATALOGUE", eyebrow_style),
            Paragraph("WHOLESALE COLLECTION", cover_title_style),
            Paragraph(
                "A curated, buyer-ready collection with clear product references, "
                "wholesale pricing, and suggested retail pricing.",
                cover_subtitle_style,
            ),
        ]
    )
    cover_metrics = Table(
        [
            [
                Paragraph(
                    f"<font color='#E1BC6A' size='20'><b>{len(products):02d}</b></font><br/>"
                    "<font color='#E8D7AF' size='6.5'>PRODUCT SKUS</font>",
                    ParagraphStyle("CoverMetric", parent=styles["Normal"], alignment=TA_CENTER),
                ),
                Paragraph(
                    f"<font color='#E1BC6A' size='20'><b>{len(grouped):02d}</b></font><br/>"
                    "<font color='#E8D7AF' size='6.5'>COLLECTIONS</font>",
                    ParagraphStyle("CoverMetric2", parent=styles["Normal"], alignment=TA_CENTER),
                ),
                Paragraph(
                    f"<font color='#E1BC6A' size='10'><b>{datetime.now():%B %Y}</b></font><br/>"
                    "<font color='#E8D7AF' size='6.5'>CURRENT EDITION</font>",
                    ParagraphStyle("CoverMetric3", parent=styles["Normal"], alignment=TA_CENTER),
                ),
            ]
        ],
        colWidths=[document.width / 3] * 3,
        rowHeights=[0.68 * inch],
    )
    cover_metrics.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), charcoal),
                ("BOX", (0, 0), (-1, -1), 0.65, gold),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#5B4825")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    story.extend(
        [
            cover_metrics,
            Spacer(1, 0.18 * inch),
            Paragraph(
                "TRADE PRICING  |  USD",
                ParagraphStyle(
                    "CatalogPricingNoteLabel",
                    parent=eyebrow_style,
                    fontSize=7,
                    textColor=bright_gold,
                    spaceAfter=4,
                ),
            ),
            Paragraph(
                "Wholesale is the trade unit price. MSRP is the suggested retail price. "
                "All amounts are shown in US dollars.",
                ParagraphStyle(
                    "CatalogPricingNote",
                    parent=styles["Normal"],
                    fontSize=7.8,
                    leading=11,
                    textColor=champagne,
                    alignment=TA_CENTER,
                ),
            ),
            PageBreak(),
        ]
    )

    card_gap = 0.12 * inch
    card_width = (document.width - (card_gap * 2)) / 3

    def product_card(product: Product) -> Table:
        image_size = card_width - 0.2 * inch
        thumbnail = Image(
            _catalog_thumbnail(product.image_url, upload_dir),
            width=image_size,
            height=image_size,
        )
        image_frame = Table(
            [[thumbnail]],
            colWidths=[image_size],
            rowHeights=[image_size],
        )
        image_frame.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), paper),
                    ("BOX", (0, 0), (-1, -1), 0.55, line),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
        right_price_label_style = ParagraphStyle(
            "CatalogRightPriceLabel", parent=price_label_style, alignment=TA_RIGHT
        )
        price_table = Table(
            [
                [
                    Paragraph("WHOLESALE", price_label_style),
                    Paragraph("MSRP", right_price_label_style),
                ],
                [
                    Paragraph(_usd(product.cost_price), wholesale_style),
                    Paragraph(_usd(product.selling_price), msrp_style),
                ],
            ],
            colWidths=[(card_width - 0.16 * inch) / 2] * 2,
        )
        price_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), obsidian),
                    ("LEFTPADDING", (0, 0), (0, -1), 8),
                    ("RIGHTPADDING", (1, 0), (1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, 0), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 0),
                    ("TOPPADDING", (0, 1), (-1, 1), 0),
                    ("BOTTOMPADDING", (0, 1), (-1, 1), 7),
                ]
            )
        )
        card = Table(
            [
                [image_frame],
                [Paragraph(escape(_shorten(product.name, 68)), name_style)],
                [Paragraph(f"SKU  {escape(_shorten(product.article_no, 38))}", sku_style)],
                [price_table],
            ],
            colWidths=[card_width],
        )
        card.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 2), paper),
                    ("BOX", (0, 0), (-1, -1), 0.6, line),
                    ("LINEABOVE", (0, 0), (-1, 0), 2.1, gold),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, 2), 7),
                    ("RIGHTPADDING", (0, 0), (-1, 2), 7),
                    ("LEFTPADDING", (0, 3), (-1, 3), 0),
                    ("RIGHTPADDING", (0, 3), (-1, 3), 0),
                    ("TOPPADDING", (0, 0), (-1, 0), 7),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
                    ("TOPPADDING", (0, 1), (-1, 1), 5),
                    ("BOTTOMPADDING", (0, 1), (-1, 1), 2),
                    ("TOPPADDING", (0, 2), (-1, 2), 1),
                    ("BOTTOMPADDING", (0, 2), (-1, 2), 6),
                    ("TOPPADDING", (0, 3), (-1, 3), 0),
                    ("BOTTOMPADDING", (0, 3), (-1, 3), 0),
                ]
            )
        )
        return card

    for category_index, (category, category_products) in enumerate(grouped.items(), start=1):
        cards = [product_card(product) for product in category_products]
        grid_widths = [card_width, card_gap, card_width, card_gap, card_width]
        card_chunks = [cards[offset : offset + 6] for offset in range(0, len(cards), 6)]
        for chunk_index, chunk in enumerate(card_chunks):
            if category_index > 1 or chunk_index > 0:
                story.append(PageBreak())

            continuation = ""
            if len(card_chunks) > 1:
                continuation = f"  |  {chunk_index + 1}/{len(card_chunks)}"
            category_header = Table(
                [
                    [
                        Paragraph(f"{category_index:02d}", category_number_style),
                        Paragraph(escape(category.upper()), category_style),
                        Paragraph(
                            f"{len(category_products)} PRODUCT{'S' if len(category_products) != 1 else ''}"
                            f"{continuation}",
                            category_count_style,
                        ),
                    ]
                ],
                colWidths=[0.55 * inch, document.width - 2.05 * inch, 1.5 * inch],
            )
            category_header.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (0, 0), gold),
                        ("BACKGROUND", (1, 0), (-1, -1), obsidian),
                        ("BOX", (0, 0), (-1, -1), 0.55, gold),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("LEFTPADDING", (0, 0), (0, 0), 0),
                        ("RIGHTPADDING", (0, 0), (0, 0), 0),
                        ("LEFTPADDING", (1, 0), (-1, -1), 11),
                        ("RIGHTPADDING", (1, 0), (-1, -1), 11),
                        ("TOPPADDING", (0, 0), (-1, -1), 7),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                    ]
                )
            )

            rows: list[list[Any]] = []
            for offset in range(0, len(chunk), 3):
                row = chunk[offset : offset + 3]
                while len(row) < 3:
                    row.append("")
                rows.append([row[0], "", row[1], "", row[2]])
            category_grid = Table(rows, colWidths=grid_widths, hAlign="LEFT")
            category_grid.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                        ("TOPPADDING", (0, 0), (-1, -1), 9),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ]
                )
            )
            story.append(KeepTogether([category_header, category_grid]))

    document.build(story, onFirstPage=page_decoration, onLaterPages=page_decoration)
    return output.getvalue()


def build_product_catalog_pdf(products: list[Product], upload_dir: Path) -> bytes:
    """Build an image-led, print-safe product catalog with modern brand styling."""
    if not products:
        raise ProductCatalogError("Add at least one product before downloading a catalog.")

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfgen import canvas as reportlab_canvas
    from reportlab.lib.utils import ImageReader
    import unicodedata

    evergreen = colors.HexColor("#123C35")
    charcoal = colors.HexColor("#22302C")
    black = colors.HexColor("#101413")
    brass = colors.HexColor("#B97828")
    warm_brass = colors.HexColor("#D19A3F")
    champagne = colors.HexColor("#EFE1C7")
    parchment = colors.HexColor("#F4EFE6")
    paper = colors.HexColor("#FFFDF8")
    cream = colors.HexColor("#F9F5ED")
    ink = colors.HexColor("#151815")
    muted = colors.HexColor("#68706D")
    hairline = colors.HexColor("#D9D1C4")
    steel = colors.HexColor("#2F4750")
    sage = colors.HexColor("#C8D3CA")
    mist = colors.HexColor("#E6ECE6")

    width, height = letter
    margin = 0.34 * inch
    content_width = width - (margin * 2)
    logo_path = Path(__file__).resolve().parent / "assets" / "hisbenew-catalog-logo.png"

    grouped: dict[str, list[Product]] = {}
    for product in sorted(
        products,
        key=lambda item: (
            (_clean_text(item.category) or "Uncategorized").casefold(),
            _clean_text(item.name).casefold(),
            _clean_text(item.article_no).casefold(),
        ),
    ):
        grouped.setdefault(_clean_text(product.category) or "Uncategorized", []).append(product)

    showcase_products = [product for product in products if _clean_text(product.image_url)] or list(products)
    hero_product = showcase_products[0]

    def safe_text(value: Any) -> str:
        normalized = unicodedata.normalize("NFKD", _clean_text(value))
        ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
        return re.sub(r"\s+", " ", ascii_text).strip()

    def fitted_text(value: Any, font_name: str, font_size: float, max_width: float) -> str:
        text = safe_text(value)
        if pdfmetrics.stringWidth(text, font_name, font_size) <= max_width:
            return text
        suffix = "..."
        while text and pdfmetrics.stringWidth(text + suffix, font_name, font_size) > max_width:
            text = text[:-1].rstrip()
        return f"{text}{suffix}" if text else suffix

    def wrapped_lines(
        value: Any,
        font_name: str,
        font_size: float,
        max_width: float,
        max_lines: int = 2,
    ) -> list[str]:
        words = safe_text(value).split()
        lines: list[str] = []
        current = ""
        while words and len(lines) < max_lines:
            word = words.pop(0)
            candidate = word if not current else f"{current} {word}"
            if pdfmetrics.stringWidth(candidate, font_name, font_size) <= max_width:
                current = candidate
                continue
            if current:
                lines.append(current)
                current = ""
                words.insert(0, word)
                continue
            lines.append(fitted_text(word, font_name, font_size, max_width))
        if current and len(lines) < max_lines:
            lines.append(current)
        if words and lines:
            remainder = f"{lines[-1]} {' '.join(words)}"
            lines[-1] = fitted_text(remainder, font_name, font_size, max_width)
        return lines or [""]

    def catalog_price(value: Any) -> str:
        try:
            amount = float(value or 0)
        except (TypeError, ValueError):
            amount = 0
        return _usd(amount) if amount > 0 else "Contact"

    thumbnail_cache: dict[str, bytes] = {}

    def image_reader(product: Product) -> ImageReader:
        cache_key = f"{getattr(product, 'id', '')}:{_clean_text(product.image_url)}"
        if cache_key not in thumbnail_cache:
            thumbnail_cache[cache_key] = _catalog_thumbnail(product.image_url, upload_dir).getvalue()
        return ImageReader(BytesIO(thumbnail_cache[cache_key]))

    def draw_catalog_image(
        product: Product,
        image_x: float,
        image_y: float,
        image_width: float,
        image_height: float,
        background_color: Any = colors.white,
        border_color: Any = hairline,
    ):
        pdf.setFillColor(background_color)
        pdf.roundRect(image_x, image_y, image_width, image_height, 6, stroke=0, fill=1)
        pdf.setStrokeColor(border_color)
        pdf.setLineWidth(0.45)
        pdf.roundRect(image_x, image_y, image_width, image_height, 6, stroke=1, fill=0)
        inset = 0.025 * inch
        pdf.drawImage(
            image_reader(product),
            image_x + inset,
            image_y + inset,
            width=image_width - (inset * 2),
            height=image_height - (inset * 2),
            preserveAspectRatio=True,
            mask="auto",
        )

    output = BytesIO()
    pdf = reportlab_canvas.Canvas(output, pagesize=letter, pageCompression=1)
    pdf.setTitle("Hisbenew Industries Product Catalog")
    pdf.setAuthor("Hisbenew Industries")
    pdf.setSubject("Image-led product catalog")

    def draw_cover():
        pdf.setFillColor(parchment)
        pdf.rect(0, 0, width, height, stroke=0, fill=1)

        sidebar_width = 0.88 * inch
        pdf.setFillColor(evergreen)
        pdf.rect(0, 0, sidebar_width, height, stroke=0, fill=1)
        pdf.setFillColor(brass)
        pdf.rect(sidebar_width, 0, 0.055 * inch, height, stroke=0, fill=1)

        if logo_path.is_file():
            logo_width = 0.56 * inch
            pdf.drawImage(
                ImageReader(str(logo_path)),
                0.16 * inch,
                height - 0.92 * inch,
                width=logo_width,
                height=logo_width / 1.5,
                preserveAspectRatio=True,
                mask="auto",
            )
        else:
            pdf.setFillColor(colors.white)
            pdf.setFont("Helvetica-Bold", 8.2)
            pdf.drawCentredString(sidebar_width / 2, height - 0.56 * inch, "HISBENEW")

        pdf.saveState()
        pdf.translate(0.40 * inch, 1.02 * inch)
        pdf.rotate(90)
        pdf.setFillColor(colors.white)
        pdf.setFont("Helvetica-Bold", 8.2)
        pdf.drawString(0, 0, "IMAGE-LED PRODUCT CATALOG")
        pdf.restoreState()

        content_x = sidebar_width + 0.38 * inch
        right_edge = width - margin
        top_y = height - 0.68 * inch
        edition = f"{len(products)} products / {len(grouped)} collections / {datetime.now():%B %Y}"

        pdf.setFillColor(brass)
        pdf.setFont("Helvetica-Bold", 7.4)
        pdf.drawString(content_x, top_y, "HISBENEW INDUSTRIES")
        pdf.setFillColor(muted)
        pdf.setFont("Helvetica", 7.1)
        pdf.drawRightString(right_edge, top_y, edition)

        hero_size = 3.82 * inch
        hero_x = right_edge - hero_size
        hero_y = 4.42 * inch
        pdf.setFillColor(mist)
        pdf.roundRect(
            hero_x - 0.12 * inch,
            hero_y - 0.12 * inch,
            hero_size + 0.24 * inch,
            hero_size + 0.24 * inch,
            10,
            stroke=0,
            fill=1,
        )
        draw_catalog_image(hero_product, hero_x, hero_y, hero_size, hero_size, colors.white, sage)

        title_width = hero_x - content_x - 0.30 * inch
        title_y = 8.78 * inch
        pdf.setFillColor(ink)
        pdf.setFont("Helvetica-Bold", 29)
        for index, line_text in enumerate(
            wrapped_lines("Modern wholesale catalog", "Helvetica-Bold", 29, title_width, max_lines=3)
        ):
            pdf.drawString(content_x, title_y - (index * 0.39 * inch), line_text.upper())

        pdf.setFillColor(steel)
        pdf.setFont("Helvetica", 9.6)
        subtitle = "Image-first product reference for handmade blades, gift sets, kitchens, and outdoor collections."
        for index, line_text in enumerate(
            wrapped_lines(subtitle, "Helvetica", 9.6, title_width, max_lines=4)
        ):
            pdf.drawString(content_x, 6.98 * inch - (index * 0.18 * inch), line_text)

        pdf.setFillColor(evergreen)
        pdf.roundRect(content_x, 5.95 * inch, 2.20 * inch, 0.32 * inch, 5, stroke=0, fill=1)
        pdf.setFillColor(colors.white)
        pdf.setFont("Helvetica-Bold", 6.8)
        pdf.drawCentredString(content_x + 1.10 * inch, 6.055 * inch, "BUYER EDITION")

        pdf.setFillColor(ink)
        pdf.setFont("Helvetica-Bold", 8)
        pdf.drawString(hero_x, hero_y - 0.22 * inch, fitted_text(hero_product.name, "Helvetica-Bold", 8, hero_size))
        pdf.setFillColor(muted)
        pdf.setFont("Helvetica", 6.3)
        pdf.drawString(
            hero_x,
            hero_y - 0.39 * inch,
            fitted_text(f"SKU {safe_text(hero_product.article_no)}", "Helvetica", 6.3, hero_size),
        )

        strip_y = 0.62 * inch
        strip_h = 1.76 * inch
        strip_w = right_edge - content_x
        pdf.setFillColor(paper)
        pdf.roundRect(content_x, strip_y, strip_w, strip_h, 8, stroke=0, fill=1)
        pdf.setStrokeColor(hairline)
        pdf.setLineWidth(0.5)
        pdf.roundRect(content_x, strip_y, strip_w, strip_h, 8, stroke=1, fill=0)
        pdf.setFillColor(evergreen)
        pdf.setFont("Helvetica-Bold", 6.2)
        pdf.drawString(content_x + 0.16 * inch, strip_y + strip_h - 0.31 * inch, "SELECTED PRODUCT VIEWS")

        thumbs = showcase_products[:4]
        thumb_gap = 0.12 * inch
        thumb_size = min(1.08 * inch, (strip_w - 0.32 * inch - (thumb_gap * 3)) / 4)
        thumb_y = strip_y + 0.31 * inch
        for index, product in enumerate(thumbs):
            thumb_x = content_x + 0.16 * inch + (index * (thumb_size + thumb_gap))
            draw_catalog_image(product, thumb_x, thumb_y, thumb_size, thumb_size, colors.white, hairline)

        pdf.setFillColor(colors.white)
        pdf.setFont("Helvetica-Bold", 6.8)
        pdf.drawCentredString(sidebar_width / 2, 0.42 * inch, "TRADE EDITION")

    def draw_page_chrome(page_number: int):
        pdf.setFillColor(parchment)
        pdf.rect(0, 0, width, height, stroke=0, fill=1)
        pdf.setFillColor(evergreen)
        pdf.rect(0, height - 0.38 * inch, width, 0.38 * inch, stroke=0, fill=1)
        pdf.setFillColor(brass)
        pdf.rect(0, height - 0.405 * inch, width, 0.025 * inch, stroke=0, fill=1)
        pdf.setFont("Helvetica-Bold", 7.4)
        pdf.setFillColor(colors.white)
        pdf.drawString(margin, height - 0.245 * inch, "HISBENEW")
        pdf.setFont("Helvetica", 7.0)
        pdf.setFillColor(champagne)
        pdf.drawRightString(width - margin, height - 0.245 * inch, "PRODUCT CATALOG")

        pdf.setStrokeColor(hairline)
        pdf.setLineWidth(0.4)
        pdf.line(margin, 0.265 * inch, width - margin, 0.265 * inch)
        pdf.setFont("Helvetica", 5.9)
        pdf.setFillColor(muted)
        pdf.drawString(margin, 0.13 * inch, "Prices shown in USD. Confirm current availability before order confirmation.")
        pdf.setFont("Helvetica-Bold", 6.5)
        pdf.setFillColor(ink)
        pdf.drawRightString(width - margin, 0.13 * inch, f"{page_number:02d}")

    def draw_category_header(
        category_index: int,
        category: str,
        product_count: int,
        banner_y: float,
    ):
        banner_height = 0.30 * inch
        pdf.setFillColor(paper)
        pdf.roundRect(margin, banner_y, content_width, banner_height, 4, stroke=0, fill=1)
        pdf.setStrokeColor(hairline)
        pdf.setLineWidth(0.45)
        pdf.roundRect(margin, banner_y, content_width, banner_height, 4, stroke=1, fill=0)
        pdf.setFillColor(evergreen)
        pdf.rect(margin, banner_y, 0.055 * inch, banner_height, stroke=0, fill=1)
        pdf.setFillColor(brass)
        pdf.setFont("Helvetica-Bold", 6.4)
        pdf.drawString(margin + 0.14 * inch, banner_y + 0.105 * inch, f"{category_index:02d}")
        pdf.setFillColor(ink)
        pdf.setFont("Helvetica-Bold", 9.3)
        category_label = fitted_text(category.upper(), "Helvetica-Bold", 9.3, 4.60 * inch)
        pdf.drawString(margin + 0.47 * inch, banner_y + 0.103 * inch, category_label)
        count_label = f"{product_count} PRODUCT{'S' if product_count != 1 else ''}"
        pdf.setFillColor(muted)
        pdf.setFont("Helvetica", 6.2)
        pdf.drawRightString(width - margin - 0.12 * inch, banner_y + 0.108 * inch, count_label)

    columns = 4
    card_gap = 0.085 * inch
    card_width = (content_width - (card_gap * (columns - 1))) / columns
    card_height = 2.17 * inch
    row_gap = 0.035 * inch
    image_size = min(card_width - 0.16 * inch, 1.62 * inch)

    def draw_product_card(product: Product, column: int, card_y: float):
        card_x = margin + (column * (card_width + card_gap))
        pdf.setFillColor(paper)
        pdf.roundRect(card_x, card_y, card_width, card_height, 5, stroke=0, fill=1)
        pdf.setStrokeColor(hairline)
        pdf.setLineWidth(0.45)
        pdf.roundRect(card_x, card_y, card_width, card_height, 5, stroke=1, fill=0)
        pdf.setFillColor(warm_brass)
        pdf.rect(card_x + 0.08 * inch, card_y + card_height - 0.032 * inch, card_width - 0.16 * inch, 0.016 * inch, stroke=0, fill=1)

        image_x = card_x + ((card_width - image_size) / 2)
        image_y = card_y + card_height - 0.080 * inch - image_size
        draw_catalog_image(product, image_x, image_y, image_size, image_size, colors.white, hairline)

        text_x = card_x + 0.08 * inch
        text_width = card_width - 0.16 * inch
        name_top = image_y - 0.065 * inch
        name_lines = wrapped_lines(product.name, "Helvetica-Bold", 6.2, text_width, max_lines=2)
        pdf.setFillColor(ink)
        pdf.setFont("Helvetica-Bold", 6.2)
        for index, line_text in enumerate(name_lines):
            pdf.drawString(text_x, name_top - (index * 0.075 * inch), line_text)

        pdf.setStrokeColor(hairline)
        pdf.setLineWidth(0.35)
        pdf.line(text_x, card_y + 0.285 * inch, card_x + card_width - 0.08 * inch, card_y + 0.285 * inch)

        pdf.setFillColor(muted)
        pdf.setFont("Helvetica", 5.1)
        sku = fitted_text(f"SKU {safe_text(product.article_no)}", "Helvetica", 5.1, text_width)
        pdf.drawString(text_x, card_y + 0.172 * inch, sku)

        price_line = fitted_text(
            f"W {catalog_price(product.cost_price)} / MSRP {catalog_price(product.selling_price)}",
            "Helvetica",
            5.1,
            text_width,
        )
        pdf.setFillColor(steel)
        pdf.setFont("Helvetica", 5.1)
        pdf.drawString(text_x, card_y + 0.060 * inch, price_line)

    draw_cover()
    pdf.showPage()
    page_number = 2
    draw_page_chrome(page_number)
    cursor_y = height - 0.52 * inch
    bottom_limit = 0.26 * inch
    category_header_height = 0.30 * inch
    category_gap = 0.025 * inch
    header_to_cards = 0.055 * inch

    for category_index, (category, category_products) in enumerate(grouped.items(), start=1):
        required = category_header_height + header_to_cards + card_height
        if cursor_y - required < bottom_limit:
            pdf.showPage()
            page_number += 1
            draw_page_chrome(page_number)
            cursor_y = height - 0.52 * inch

        banner_y = cursor_y - category_header_height
        draw_category_header(category_index, category, len(category_products), banner_y)
        cursor_y = banner_y - header_to_cards

        rows = [category_products[offset : offset + columns] for offset in range(0, len(category_products), columns)]
        for row in rows:
            if cursor_y - card_height < bottom_limit:
                pdf.showPage()
                page_number += 1
                draw_page_chrome(page_number)
                cursor_y = height - 0.52 * inch
            card_y = cursor_y - card_height
            for column, product in enumerate(row):
                draw_product_card(product, column, card_y)
            cursor_y = card_y - row_gap
        cursor_y -= category_gap

    pdf.save()
    return output.getvalue()